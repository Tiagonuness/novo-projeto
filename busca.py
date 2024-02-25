import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import openpyxl

driver = webdriver.Chrome()

url = "https://www.dfimoveis.com.br/aluguel/df/brasilia/apartamento?&ordenamento=mais-recente"
driver.get(url)

arquivoExcel = openpyxl.load_workbook('Apartamentos.xlsx')
planilha = arquivoExcel['Apartamentos']

tamanhoDoResultadoDeBusca = len(driver.find_element(By.ID, 'resultadoDaBuscaDeImoveis').find_elements(By.TAG_NAME, 'a'))

for apartamento in range(tamanhoDoResultadoDeBusca):
    
    #Adcionar a ordem do imóvel no arquivo Excel
    linhaComeco = (1 + apartamento) + 1
    planilha.cell(row=linhaComeco, column=1).value = apartamento + 1

    #Coletar e adcionar o endereço no arquivo Excel
    endereço = driver.find_elements(By.TAG_NAME, 'h2')[apartamento].text
    planilha.cell(row=linhaComeco, column=2).value = endereço

    #Coletar e adcionar o valor do imóvel no arquivo Excel
    valor = driver.find_element(By.ID, 'resultadoDaBuscaDeImoveis').find_elements(By.TAG_NAME, 'a')[apartamento].find_elements(By.TAG_NAME, 'h4')[0]
    valor = valor.find_element(By.TAG_NAME, 'span').text
    valor = 'R$' + ' ' + valor
    planilha.cell(row=linhaComeco, column=3).value = valor

    #Coletar e adcionar o valor por metro quadrado do imóvel no arquivo Excel
    valorm2 = driver.find_element(By.ID, 'resultadoDaBuscaDeImoveis').find_elements(By.TAG_NAME, 'a')[apartamento].find_elements(By.TAG_NAME, 'h4')[1]
    valorm2 = valorm2.find_element(By.TAG_NAME, 'span').text
    valorm2 = 'R$' + ' ' + valorm2
    planilha.cell(row=linhaComeco, column=4).value = valorm2

    #Coletar a área,quartos e suites do imóvel e adcionar na planilha Excel
    imovel = driver.find_element(By.ID, 'resultadoDaBuscaDeImoveis').find_elements(By.TAG_NAME, 'a')[apartamento].find_elements(By.CLASS_NAME, 'new-details-ul')[0].text
    detalhes = imovel.split('\n')
    for li in range(len(detalhes)):
        if 'm²' in detalhes[li]: planilha.cell(row=linhaComeco, column=5).value = detalhes[li]
        elif 'Quartos' in detalhes[li] or 'Quarto' in detalhes[li]: planilha.cell(row=linhaComeco, column=6).value = detalhes[li]
        elif 'Suítes' in detalhes[li] or 'Suíte' in detalhes[li]: planilha.cell(row=linhaComeco, column=7).value = detalhes[li]
        elif 'Vagas' in detalhes[li] or 'Vaga' in detalhes[li]: planilha.cell(row=linhaComeco, column=8).value = detalhes[li]

    #Obter link do imóvel e adcionar na planilha Excel
    imoveLink = driver.find_element(By.ID, 'resultadoDaBuscaDeImoveis').find_elements(By.TAG_NAME, 'a')[apartamento].get_attribute('href')
    planilha.cell(row=linhaComeco, column=9).value = imoveLink

    #Salvar todas alterações 
    arquivoExcel.save('Apartamentos.xlsx')